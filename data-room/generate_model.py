#!/usr/bin/env python3
"""
CacheCow / ForcedField Technologies — 5-Year Financial Model Generator
Generates an interactive .xlsx workbook with Excel formulas throughout.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, AreaChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart.layout import Layout, ManualLayout

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "cachecow_financial_model.xlsx")

# ── Colour palette ──────────────────────────────────────────────────
DK_GREEN = "1a3d2b"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
BLUE_TAB = "4472C4"
ORANGE_TAB = "ED7D31"
RED_TAB = "C00000"
PURPLE_TAB = "7030A0"
TEAL_TAB = "00B0F0"
YELLOW_TAB = "FFC000"
GREEN_TAB = "00B050"
GREY_TAB = "808080"

hdr_fill = PatternFill("solid", fgColor=DK_GREEN)
hdr_font = Font(bold=True, color=WHITE, size=11)
section_font = Font(bold=True, size=11)
pct_fmt = '0.0%'
cur_fmt = '$#,##0'
cur_fmt_neg = '$#,##0;($#,##0)'
num_fmt = '#,##0'
cur_fmt_k = '$#,##0'

thin_border = Border(
    bottom=Side(style='thin', color='D9D9D9')
)


def style_header_row(ws, row, max_col=6):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def set_col_widths(ws, a_width=35, other_width=18, max_col=6):
    ws.column_dimensions['A'].width = a_width
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = other_width


def freeze(ws, cell='A3'):
    ws.freeze_panes = cell


def fmt_row(ws, row, fmt, cols=range(2, 7)):
    for c in cols:
        ws.cell(row=row, column=c).number_format = fmt


# ════════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Sheet 1: Cover ─────────────────────────────────────────────────
ws_cover = wb.active
ws_cover.title = "Cover"
ws_cover.sheet_properties.tabColor = GREEN_TAB

ws_cover['A1'] = "CacheCow / ForcedField Technologies"
ws_cover['A1'].font = Font(bold=True, size=18, color=DK_GREEN)
ws_cover['A2'] = "5-Year Financial Model"
ws_cover['A2'].font = Font(size=14)
ws_cover['A3'] = "Confidential \u2014 April 2026"
ws_cover['A3'].font = Font(size=11, italic=True)

ws_cover['A5'] = "Scenario:"
ws_cover['A5'].font = Font(bold=True)
ws_cover['B5'] = "Base"
dv = DataValidation(type="list", formula1='"Conservative,Base,Optimistic"', allow_blank=False)
dv.error = "Please select a valid scenario"
dv.errorTitle = "Invalid Scenario"
ws_cover.add_data_validation(dv)
dv.add(ws_cover['B5'])

ws_cover['A7'] = "Table of Contents"
ws_cover['A7'].font = Font(bold=True, size=12)
sheets_list = [
    ("1", "Cover", "Overview and scenario selector"),
    ("2", "Assumptions", "All model inputs and parameters"),
    ("3", "Revenue Buildup", "Revenue by stream, Y1-Y5"),
    ("4", "P&L", "Profit & Loss statement"),
    ("5", "Unit Economics", "Per-device and per-ranch metrics"),
    ("6", "Scenarios", "Conservative / Base / Optimistic comparison"),
    ("7", "Cash Flow", "Operating cash flow and runway"),
    ("8", "Cap Table", "Ownership through Seed, A, B"),
    ("9", "Charts", "Visual dashboards"),
]
for i, (num, name, desc) in enumerate(sheets_list):
    r = 8 + i
    ws_cover[f'A{r}'] = f"  {num}. {name}"
    ws_cover[f'B{r}'] = desc

ws_cover.column_dimensions['A'].width = 35
ws_cover.column_dimensions['B'].width = 50

# ── Sheet 2: Assumptions ──────────────────────────────────────────
ws_a = wb.create_sheet("Assumptions")
ws_a.sheet_properties.tabColor = BLUE_TAB
set_col_widths(ws_a)

# Helper
def a_hdr(row, title):
    ws_a.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color=DK_GREEN)
    style_header_row(ws_a, row, 6)
    ws_a.cell(row=row, column=1).fill = hdr_fill
    ws_a.cell(row=row, column=1).font = Font(bold=True, size=12, color=WHITE)

# --- Device & Pricing ---
a_hdr(2, "Device & Pricing")
labels_dp = [
    ("Device Unit Price ($)", 95, cur_fmt),
    ("Annual Subscription ($/animal/yr)", 50, cur_fmt),
    ("Hardware COGS per Device ($)", 33, cur_fmt),
    ("Subscription COGS per Animal/yr ($)", 17.50, cur_fmt),
    ("Hardware Gross Margin", None, pct_fmt),
    ("Subscription Gross Margin", None, pct_fmt),
]
for i, (lbl, val, fmt) in enumerate(labels_dp):
    r = 3 + i
    ws_a.cell(row=r, column=1, value=lbl)
    if val is not None:
        ws_a.cell(row=r, column=2, value=val).number_format = fmt
    else:
        ws_a.cell(row=r, column=2).number_format = fmt

ws_a['B7'] = '=(B3-B5)/B3'
ws_a['B7'].number_format = pct_fmt
ws_a['B8'] = '=(B4-B6)/B4'
ws_a['B8'].number_format = pct_fmt

# --- Deployment Schedule ---
a_hdr(10, "Deployment Schedule")
ws_a.cell(row=11, column=1, value="")
yr_labels = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
for i, lbl in enumerate(yr_labels):
    ws_a.cell(row=11, column=2+i, value=lbl).font = Font(bold=True)
    ws_a.cell(row=11, column=2+i).alignment = Alignment(horizontal='center')

ws_a.cell(row=12, column=1, value="Devices Deployed (Cumulative)")
cum_devices = [10000, 200000, 500000, 750000, 1000000]
for i, v in enumerate(cum_devices):
    ws_a.cell(row=12, column=2+i, value=v).number_format = num_fmt

ws_a.cell(row=13, column=1, value="Net New Devices")
new_devices = [10000, 190000, 300000, 250000, 250000]
for i, v in enumerate(new_devices):
    ws_a.cell(row=13, column=2+i, value=v).number_format = num_fmt

ws_a.cell(row=14, column=1, value="Annual Churn Rate")
for i in range(5):
    ws_a.cell(row=14, column=2+i, value=0.02).number_format = pct_fmt

ws_a.cell(row=15, column=1, value="Active Devices (End of Year)")
# Y1: just net new (no prior year)
ws_a['B15'] = '=B13'
ws_a['B15'].number_format = num_fmt
# Y2-Y5: prior active * (1-churn) + new
for i in range(1, 5):
    col = get_column_letter(2 + i)
    prev = get_column_letter(1 + i)
    ws_a[f'{col}15'] = f'={prev}15*(1-{col}14)+{col}13'
    ws_a[f'{col}15'].number_format = num_fmt

# --- Revenue Layer Activation ---
a_hdr(17, "Revenue Layer Activation")
ws_a.cell(row=18, column=1, value="")
for i, lbl in enumerate(yr_labels):
    ws_a.cell(row=18, column=2+i, value=lbl).font = Font(bold=True)

ws_a.cell(row=19, column=1, value="Hardware + Subscription")
for i in range(5):
    ws_a.cell(row=19, column=2+i, value=1)

ws_a.cell(row=20, column=1, value="Data Licensing")
dl_act = [0, 1, 1, 1, 1]
for i, v in enumerate(dl_act):
    ws_a.cell(row=20, column=2+i, value=v)

ws_a.cell(row=21, column=1, value="Platform IP")
ip_act = [0, 0, 1, 1, 1]
for i, v in enumerate(ip_act):
    ws_a.cell(row=21, column=2+i, value=v)

# --- Weather Participation Rates ---
a_hdr(23, "Weather Participation Rates")
ws_a.cell(row=24, column=1, value="Conservative")
ws_a['B24'] = 0.10
ws_a['B24'].number_format = pct_fmt
ws_a.cell(row=25, column=1, value="Base")
ws_a['B25'] = 0.20
ws_a['B25'].number_format = pct_fmt
ws_a.cell(row=26, column=1, value="Optimistic")
ws_a['B26'] = 0.30
ws_a['B26'].number_format = pct_fmt

# --- Data Licensing ---
a_hdr(28, "Data Licensing")
ws_a.cell(row=29, column=1, value="Addressable TAM at Scale ($)")
ws_a['B29'] = 60000000
ws_a['B29'].number_format = cur_fmt
ws_a.cell(row=30, column=1, value="Data Licensing per Device/yr (Base, $)")
ws_a['B30'] = 15
ws_a['B30'].number_format = cur_fmt
ws_a.cell(row=31, column=1, value="Revenue Share to Ranchers")
ws_a['B31'] = 0.15
ws_a['B31'].number_format = pct_fmt

# --- Platform IP Licensing ---
a_hdr(33, "Platform IP Licensing")
ws_a.cell(row=34, column=1, value="IP License Deals Y3")
ws_a['B34'] = 2
ws_a.cell(row=35, column=1, value="IP License Deals Y4")
ws_a['B35'] = 5
ws_a.cell(row=36, column=1, value="IP License Deals Y5")
ws_a['B36'] = 12

ws_a.cell(row=37, column=1, value="Avg Deal Size ($)")
deal_sizes = [0, 0, 500000, 1000000, 2000000]
for i, v in enumerate(deal_sizes):
    ws_a.cell(row=37, column=2+i, value=v).number_format = cur_fmt

# IP deals row for reference in formulas (Y1-Y5)
ws_a.cell(row=38, column=1, value="IP Deals by Year")
ip_deals = [0, 0, 2, 5, 12]
for i, v in enumerate(ip_deals):
    ws_a.cell(row=38, column=2+i, value=v)

# --- OpEx ---
a_hdr(39, "Operating Expenses")
ws_a.cell(row=40, column=1, value="")
for i, lbl in enumerate(yr_labels):
    ws_a.cell(row=40, column=2+i, value=lbl).font = Font(bold=True)

ws_a.cell(row=41, column=1, value="Headcount")
hc = [5, 12, 30, 60, 100]
for i, v in enumerate(hc):
    ws_a.cell(row=41, column=2+i, value=v).number_format = num_fmt

ws_a.cell(row=42, column=1, value="Avg Salary ($)")
salaries = [120000, 126000, 132300, 138915, 145861]
for i, v in enumerate(salaries):
    ws_a.cell(row=42, column=2+i, value=v).number_format = cur_fmt

ws_a.cell(row=43, column=1, value="Total People Cost ($)")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_a[f'{col}43'] = f'={col}41*{col}42*1.3'
    ws_a[f'{col}43'].number_format = cur_fmt

ws_a.cell(row=44, column=1, value="R&D % of Revenue")
rd_pct = [0.40, 0.35, 0.30, 0.27, 0.25]
for i, v in enumerate(rd_pct):
    ws_a.cell(row=44, column=2+i, value=v).number_format = pct_fmt

ws_a.cell(row=45, column=1, value="S&M % of Revenue")
sm_pct = [0.15, 0.18, 0.20, 0.18, 0.15]
for i, v in enumerate(sm_pct):
    ws_a.cell(row=45, column=2+i, value=v).number_format = pct_fmt

ws_a.cell(row=46, column=1, value="G&A % of Revenue")
ga_pct = [0.15, 0.12, 0.10, 0.10, 0.10]
for i, v in enumerate(ga_pct):
    ws_a.cell(row=46, column=2+i, value=v).number_format = pct_fmt

# --- Fundraising ---
a_hdr(49, "Fundraising")
ws_a.cell(row=50, column=1, value="Round")
ws_a.cell(row=50, column=2, value="Amount ($)")
ws_a.cell(row=50, column=3, value="Pre-money ($)")
ws_a.cell(row=50, column=4, value="Post-money ($)")
ws_a.cell(row=50, column=5, value="Timing")
for c in range(1, 6):
    ws_a.cell(row=50, column=c).font = Font(bold=True)

ws_a.cell(row=51, column=1, value="Seed")
ws_a['B51'] = 3500000; ws_a['B51'].number_format = cur_fmt
ws_a['C51'] = 15000000; ws_a['C51'].number_format = cur_fmt
ws_a['D51'] = '=B51+C51'; ws_a['D51'].number_format = cur_fmt
ws_a['E51'] = "Y1"

ws_a.cell(row=52, column=1, value="Series A")
ws_a['B52'] = 15000000; ws_a['B52'].number_format = cur_fmt
ws_a['C52'] = 75000000; ws_a['C52'].number_format = cur_fmt
ws_a['D52'] = '=B52+C52'; ws_a['D52'].number_format = cur_fmt
ws_a['E52'] = "Y2"

ws_a.cell(row=53, column=1, value="Series B")
ws_a['B53'] = 75000000; ws_a['B53'].number_format = cur_fmt
ws_a['C53'] = 400000000; ws_a['C53'].number_format = cur_fmt
ws_a['D53'] = '=B53+C53'; ws_a['D53'].number_format = cur_fmt
ws_a['E53'] = "Y3"

# --- Margins by Stream ---
a_hdr(57, "Margins by Stream")
ws_a.cell(row=58, column=1, value="Hardware + Sub Gross Margin")
ws_a['B58'] = 0.65; ws_a['B58'].number_format = pct_fmt
ws_a.cell(row=59, column=1, value="Data Licensing Gross Margin")
ws_a['B59'] = 0.95; ws_a['B59'].number_format = pct_fmt
ws_a.cell(row=60, column=1, value="Platform IP Gross Margin")
ws_a['B60'] = 0.90; ws_a['B60'].number_format = pct_fmt
ws_a.cell(row=61, column=1, value="Weather Data Gross Margin")
ws_a['B61'] = 0.88; ws_a['B61'].number_format = pct_fmt

freeze(ws_a, 'A3')

# ── Sheet 3: Revenue Buildup ──────────────────────────────────────
ws_r = wb.create_sheet("Revenue Buildup")
ws_r.sheet_properties.tabColor = ORANGE_TAB
set_col_widths(ws_r)

# Year headers
ws_r.cell(row=1, column=1, value="Revenue Buildup")
ws_r['A1'].font = Font(bold=True, size=14, color=DK_GREEN)
style_header_row(ws_r, 2)
ws_r.cell(row=2, column=1, value="")
for i, lbl in enumerate(yr_labels):
    ws_r.cell(row=2, column=2+i, value=lbl)

# Section 1: Hardware + Subscription
ws_r.cell(row=3, column=1, value="HARDWARE + SUBSCRIPTION").font = section_font
style_header_row(ws_r, 3)

ws_r.cell(row=4, column=1, value="Hardware Revenue")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}4'] = f'=Assumptions!{col}13*Assumptions!B3*Assumptions!{col}19'
    ws_r[f'{col}4'].number_format = cur_fmt

ws_r.cell(row=5, column=1, value="Subscription Revenue")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}5'] = f'=Assumptions!{col}15*Assumptions!B4*Assumptions!{col}19'
    ws_r[f'{col}5'].number_format = cur_fmt

ws_r.cell(row=6, column=1, value="Total H+S Revenue")
ws_r['A6'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}6'] = f'={col}4+{col}5'
    ws_r[f'{col}6'].number_format = cur_fmt

# Section 2: Data Licensing
ws_r.cell(row=8, column=1, value="").font = section_font
ws_r.cell(row=9, column=1, value="DATA LICENSING").font = section_font
style_header_row(ws_r, 9)

ws_r.cell(row=10, column=1, value="")
ws_r.cell(row=11, column=1, value="Eligible Devices")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}11'] = f'=Assumptions!{col}15*Assumptions!{col}20'
    ws_r[f'{col}11'].number_format = num_fmt

ws_r.cell(row=12, column=1, value="Data Revenue (Gross)")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}12'] = f'={col}11*Assumptions!B30'
    ws_r[f'{col}12'].number_format = cur_fmt

ws_r.cell(row=13, column=1, value="Rancher Revenue Share")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}13'] = f'={col}12*Assumptions!B31'
    ws_r[f'{col}13'].number_format = cur_fmt

ws_r.cell(row=14, column=1, value="Net Data Revenue")
ws_r['A14'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}14'] = f'={col}12-{col}13'
    ws_r[f'{col}14'].number_format = cur_fmt

# Section 3: Weather Data
ws_r.cell(row=16, column=1, value="WEATHER DATA").font = section_font
style_header_row(ws_r, 16)

ws_r.cell(row=17, column=1, value="Weather-Active Devices")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}17'] = f'=Assumptions!{col}15*Assumptions!B25'
    ws_r[f'{col}17'].number_format = num_fmt

ws_r.cell(row=18, column=1, value="Weather License Revenue")
weather_rev = [0, 8000000, 21500000, 35000000, 50000000]
for i, v in enumerate(weather_rev):
    ws_r.cell(row=18, column=2+i, value=v).number_format = cur_fmt

ws_r.cell(row=19, column=1, value="Ground Truth Premium")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}19'] = f'={col}18*0.25'
    ws_r[f'{col}19'].number_format = cur_fmt

ws_r.cell(row=20, column=1, value="Total Weather Revenue")
ws_r['A20'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}20'] = f'={col}18+{col}19'
    ws_r[f'{col}20'].number_format = cur_fmt

# Section 4: Platform IP
ws_r.cell(row=22, column=1, value="PLATFORM IP LICENSING").font = section_font
style_header_row(ws_r, 22)

ws_r.cell(row=23, column=1, value="License Deals")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}23'] = f'=Assumptions!{col}38'
    ws_r[f'{col}23'].number_format = num_fmt

ws_r.cell(row=24, column=1, value="Platform IP Revenue")
ws_r['A24'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}24'] = f'={col}23*Assumptions!{col}37*Assumptions!{col}21'
    ws_r[f'{col}24'].number_format = cur_fmt

# TOTAL REVENUE
ws_r.cell(row=26, column=1, value="")
ws_r.cell(row=27, column=1, value="")
ws_r.cell(row=28, column=1, value="TOTAL REVENUE")
ws_r['A28'].font = Font(bold=True, size=12, color=DK_GREEN)
style_header_row(ws_r, 28)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_r[f'{col}28'] = f'={col}6+{col}14+{col}20+{col}24'
    ws_r[f'{col}28'].number_format = cur_fmt
    ws_r[f'{col}28'].font = Font(bold=True, color=WHITE)

freeze(ws_r)

# ── Sheet 4: P&L ──────────────────────────────────────────────────
ws_p = wb.create_sheet("P&L")
ws_p.sheet_properties.tabColor = RED_TAB
set_col_widths(ws_p)

ws_p.cell(row=1, column=1, value="Profit & Loss").font = Font(bold=True, size=14, color=DK_GREEN)
style_header_row(ws_p, 2)
ws_p.cell(row=2, column=1, value="")
for i, lbl in enumerate(yr_labels):
    ws_p.cell(row=2, column=2+i, value=lbl)

ws_p.cell(row=3, column=1, value="Total Revenue")
ws_p['A3'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}3'] = f"='Revenue Buildup'!{col}28"
    ws_p[f'{col}3'].number_format = cur_fmt

ws_p.cell(row=4, column=1, value="").font = section_font
ws_p.cell(row=5, column=1, value="Hardware COGS")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}5'] = f'=Assumptions!{col}13*Assumptions!B5'
    ws_p[f'{col}5'].number_format = cur_fmt

ws_p.cell(row=6, column=1, value="Subscription COGS")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}6'] = f'=Assumptions!{col}15*Assumptions!B6'
    ws_p[f'{col}6'].number_format = cur_fmt

ws_p.cell(row=7, column=1, value="Data COGS")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}7'] = f"='Revenue Buildup'!{col}14*(1-Assumptions!B59)"
    ws_p[f'{col}7'].number_format = cur_fmt

ws_p.cell(row=8, column=1, value="Weather COGS")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}8'] = f"='Revenue Buildup'!{col}20*(1-Assumptions!B61)"
    ws_p[f'{col}8'].number_format = cur_fmt

ws_p.cell(row=9, column=1, value="Platform COGS")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}9'] = f"='Revenue Buildup'!{col}24*(1-Assumptions!B60)"
    ws_p[f'{col}9'].number_format = cur_fmt

ws_p.cell(row=10, column=1, value="Total COGS")
ws_p['A10'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}10'] = f'={col}5+{col}6+{col}7+{col}8+{col}9'
    ws_p[f'{col}10'].number_format = cur_fmt

ws_p.cell(row=11, column=1, value="")
ws_p.cell(row=12, column=1, value="Gross Profit")
ws_p['A12'].font = Font(bold=True, size=11, color=DK_GREEN)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}12'] = f'={col}3-{col}10'
    ws_p[f'{col}12'].number_format = cur_fmt

ws_p.cell(row=13, column=1, value="Gross Margin %")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}13'] = f'=IF({col}3=0,0,{col}12/{col}3)'
    ws_p[f'{col}13'].number_format = pct_fmt

ws_p.cell(row=14, column=1, value="")
ws_p.cell(row=15, column=1, value="R&D Expense")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}15'] = f'=MAX({col}3*Assumptions!{col}44, Assumptions!{col}43*0.6)'
    ws_p[f'{col}15'].number_format = cur_fmt

ws_p.cell(row=16, column=1, value="Sales & Marketing")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}16'] = f'={col}3*Assumptions!{col}45'
    ws_p[f'{col}16'].number_format = cur_fmt

ws_p.cell(row=17, column=1, value="G&A Expense")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}17'] = f'={col}3*Assumptions!{col}46'
    ws_p[f'{col}17'].number_format = cur_fmt

ws_p.cell(row=18, column=1, value="Total OpEx")
ws_p['A18'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}18'] = f'={col}15+{col}16+{col}17'
    ws_p[f'{col}18'].number_format = cur_fmt

ws_p.cell(row=19, column=1, value="")
ws_p.cell(row=20, column=1, value="EBITDA")
ws_p['A20'].font = Font(bold=True, size=11, color=DK_GREEN)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}20'] = f'={col}12-{col}18'
    ws_p[f'{col}20'].number_format = cur_fmt

ws_p.cell(row=21, column=1, value="EBITDA Margin %")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}21'] = f'=IF({col}3=0,0,{col}20/{col}3)'
    ws_p[f'{col}21'].number_format = pct_fmt

ws_p.cell(row=22, column=1, value="")
ws_p.cell(row=23, column=1, value="Depreciation & Amortization")
da_vals = [100000, 300000, 500000, 750000, 1000000]
for i, v in enumerate(da_vals):
    ws_p.cell(row=23, column=2+i, value=v).number_format = cur_fmt

ws_p.cell(row=24, column=1, value="Net Income")
ws_p['A24'].font = Font(bold=True, size=11, color=DK_GREEN)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_p[f'{col}24'] = f'={col}20-{col}23'
    ws_p[f'{col}24'].number_format = cur_fmt

freeze(ws_p)

# ── Sheet 5: Unit Economics ───────────────────────────────────────
ws_u = wb.create_sheet("Unit Economics")
ws_u.sheet_properties.tabColor = PURPLE_TAB
set_col_widths(ws_u)

ws_u.cell(row=1, column=1, value="Unit Economics").font = Font(bold=True, size=14, color=DK_GREEN)
style_header_row(ws_u, 2)
ws_u.cell(row=2, column=1, value="")
for i, lbl in enumerate(yr_labels):
    ws_u.cell(row=2, column=2+i, value=lbl)

rows_ue = [
    (3, "Revenue per Active Device", cur_fmt,
     lambda c: f"='P&L'!{c}3/Assumptions!{c}15"),
    (4, "COGS per Active Device", cur_fmt,
     lambda c: f"='P&L'!{c}10/Assumptions!{c}15"),
    (5, "Gross Profit per Device", cur_fmt,
     lambda c: f"={c}3-{c}4"),
    (6, "Total OpEx per Device", cur_fmt,
     lambda c: f"='P&L'!{c}18/Assumptions!{c}15"),
    (7, "Contribution Margin per Device", cur_fmt,
     lambda c: f"={c}5-{c}6"),
]
for r, label, fmt, formula_fn in rows_ue:
    ws_u.cell(row=r, column=1, value=label)
    if r in (5, 7):
        ws_u.cell(row=r, column=1).font = Font(bold=True)
    for i in range(5):
        col = get_column_letter(2 + i)
        ws_u[f'{col}{r}'] = formula_fn(col)
        ws_u[f'{col}{r}'].number_format = fmt

ws_u.cell(row=8, column=1, value="")
ws_u.cell(row=9, column=1, value="Implied Ranch Count (1000 hd avg)")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_u[f'{col}9'] = f'=Assumptions!{col}15/1000'
    ws_u[f'{col}9'].number_format = num_fmt

ws_u.cell(row=10, column=1, value="Revenue per Ranch")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_u[f'{col}10'] = f"=IF({col}9=0,0,'P&L'!{col}3/{col}9)"
    ws_u[f'{col}10'].number_format = cur_fmt

ws_u.cell(row=11, column=1, value="LTV per Ranch (10yr)")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_u[f'{col}11'] = f"={col}10*10*'P&L'!{col}13"
    ws_u[f'{col}11'].number_format = cur_fmt

ws_u.cell(row=12, column=1, value="Estimated CAC per Ranch")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_u[f'{col}12'] = f"=IF(Assumptions!{col}13=0,0,'P&L'!{col}16/(Assumptions!{col}13/1000))"
    ws_u[f'{col}12'].number_format = cur_fmt

ws_u.cell(row=13, column=1, value="LTV:CAC Ratio")
ws_u['A13'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_u[f'{col}13'] = f'=IF({col}12=0,0,{col}11/{col}12)'
    ws_u[f'{col}13'].number_format = '0.0x'

ws_u.cell(row=14, column=1, value="Payback Period (months)")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_u[f'{col}14'] = f"=IF({col}10=0,0,{col}12/(({col}10/12)*'P&L'!{col}13))"
    ws_u[f'{col}14'].number_format = '0.0'

freeze(ws_u)

# ── Sheet 6: Scenarios ────────────────────────────────────────────
ws_s = wb.create_sheet("Scenarios")
ws_s.sheet_properties.tabColor = YELLOW_TAB
set_col_widths(ws_s, a_width=30, other_width=16, max_col=16)

ws_s.cell(row=1, column=1, value="Scenario Analysis").font = Font(bold=True, size=14, color=DK_GREEN)

# Layout: Conservative B-F, Base H-L, Optimistic N-R
scenario_starts = {'Conservative': 2, 'Base': 8, 'Optimistic': 14}
scenario_factors = {'Conservative': 0.70, 'Base': 1.0, 'Optimistic': 1.30}
weather_rates = {'Conservative': 'B24', 'Base': 'B25', 'Optimistic': 'B26'}

for scenario, start_col in scenario_starts.items():
    # Header
    ws_s.cell(row=2, column=start_col, value=scenario).font = Font(bold=True, size=12, color=WHITE)
    for c in range(start_col, start_col + 5):
        ws_s.cell(row=2, column=c).fill = hdr_fill
        ws_s.cell(row=2, column=c).font = hdr_font
        ws_s.column_dimensions[get_column_letter(c)].width = 16

    # Year sub-headers
    for i, lbl in enumerate(yr_labels):
        ws_s.cell(row=3, column=start_col + i, value=lbl).font = Font(bold=True)

    # Label column (only for first scenario block)
    if start_col == 2:
        labels_sc = [
            (4, "Active Devices"),
            (5, "Total Revenue"),
            (6, "Gross Profit"),
            (7, "EBITDA"),
            (8, "Net Income"),
            (9, "Gross Margin %"),
            (10, "EBITDA Margin %"),
            (11, "Revenue per Device"),
        ]
        for r, lbl in labels_sc:
            ws_s.cell(row=r, column=1, value=lbl)
            if lbl in ("Total Revenue", "EBITDA"):
                ws_s.cell(row=r, column=1).font = Font(bold=True)

    factor = scenario_factors[scenario]
    w_rate = weather_rates[scenario]

    for i in range(5):
        col = get_column_letter(start_col + i)
        base_col = get_column_letter(2 + i)  # base assumptions column

        # Active Devices (scaled)
        ws_s[f'{col}4'] = f'=Assumptions!{base_col}15*{factor}'
        ws_s[f'{col}4'].number_format = num_fmt

        # Total Revenue (scale H+S and Data by factor, weather uses scenario rate)
        # Simplified: base revenue * factor, adjusted for weather difference
        if scenario == 'Base':
            ws_s[f'{col}5'] = f"='Revenue Buildup'!{base_col}28"
        else:
            # H+S revenue scaled
            # Data revenue scaled
            # Weather revenue scaled by weather rate ratio
            # Platform IP same for all
            ws_s[f'{col}5'] = (
                f"='Revenue Buildup'!{base_col}6*{factor}"
                f"+'Revenue Buildup'!{base_col}14*{factor}"
                f"+'Revenue Buildup'!{base_col}20*(Assumptions!{w_rate}/Assumptions!B25)"
                f"+'Revenue Buildup'!{base_col}24"
            )
        ws_s[f'{col}5'].number_format = cur_fmt

        # Gross Profit (apply blended margin from base)
        ws_s[f'{col}6'] = f"=IF('P&L'!{base_col}3=0,0,{col}5*('P&L'!{base_col}12/'P&L'!{base_col}3))"
        ws_s[f'{col}6'].number_format = cur_fmt

        # EBITDA
        ws_s[f'{col}7'] = f"=IF('P&L'!{base_col}3=0,0,{col}5*('P&L'!{base_col}20/'P&L'!{base_col}3))"
        ws_s[f'{col}7'].number_format = cur_fmt

        # Net Income
        ws_s[f'{col}8'] = f"=IF('P&L'!{base_col}3=0,0,{col}5*('P&L'!{base_col}24/'P&L'!{base_col}3))"
        ws_s[f'{col}8'].number_format = cur_fmt

        # Gross Margin %
        ws_s[f'{col}9'] = f'=IF({col}5=0,0,{col}6/{col}5)'
        ws_s[f'{col}9'].number_format = pct_fmt

        # EBITDA Margin %
        ws_s[f'{col}10'] = f'=IF({col}5=0,0,{col}7/{col}5)'
        ws_s[f'{col}10'].number_format = pct_fmt

        # Revenue per Device
        ws_s[f'{col}11'] = f'=IF({col}4=0,0,{col}5/{col}4)'
        ws_s[f'{col}11'].number_format = cur_fmt

ws_s.column_dimensions['A'].width = 22
freeze(ws_s, 'A4')

# ── Sheet 7: Cash Flow ───────────────────────────────────────────
ws_cf = wb.create_sheet("Cash Flow")
ws_cf.sheet_properties.tabColor = TEAL_TAB
set_col_widths(ws_cf)

ws_cf.cell(row=1, column=1, value="Cash Flow Statement").font = Font(bold=True, size=14, color=DK_GREEN)
style_header_row(ws_cf, 2)
ws_cf.cell(row=2, column=1, value="")
for i, lbl in enumerate(yr_labels):
    ws_cf.cell(row=2, column=2+i, value=lbl)

ws_cf.cell(row=3, column=1, value="EBITDA")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_cf[f'{col}3'] = f"='P&L'!{col}20"
    ws_cf[f'{col}3'].number_format = cur_fmt

ws_cf.cell(row=4, column=1, value="CapEx")
capex = [-500000, -2000000, -3000000, -4000000, -5000000]
for i, v in enumerate(capex):
    ws_cf.cell(row=4, column=2+i, value=v).number_format = cur_fmt

ws_cf.cell(row=5, column=1, value="Working Capital Changes")
wc = [-200000, -1000000, -1500000, -2000000, -2500000]
for i, v in enumerate(wc):
    ws_cf.cell(row=5, column=2+i, value=v).number_format = cur_fmt

ws_cf.cell(row=6, column=1, value="")
ws_cf.cell(row=7, column=1, value="Operating Cash Flow")
ws_cf['A7'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_cf[f'{col}7'] = f'={col}3+{col}4+{col}5'
    ws_cf[f'{col}7'].number_format = cur_fmt

ws_cf.cell(row=8, column=1, value="")
ws_cf.cell(row=9, column=1, value="Fundraising Inflows")
fund_inflows = [3500000, 15000000, 75000000, 0, 0]
for i, v in enumerate(fund_inflows):
    ws_cf.cell(row=9, column=2+i, value=v).number_format = cur_fmt

ws_cf.cell(row=10, column=1, value="Net Cash Flow")
ws_cf['A10'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_cf[f'{col}10'] = f'={col}7+{col}9'
    ws_cf[f'{col}10'].number_format = cur_fmt

ws_cf.cell(row=11, column=1, value="")
ws_cf.cell(row=12, column=1, value="Opening Cash")
ws_cf['B12'] = 0
ws_cf['B12'].number_format = cur_fmt
for i in range(1, 5):
    col = get_column_letter(2 + i)
    prev = get_column_letter(1 + i)
    ws_cf[f'{col}12'] = f'={prev}13'
    ws_cf[f'{col}12'].number_format = cur_fmt

ws_cf.cell(row=13, column=1, value="Closing Cash")
ws_cf['A13'].font = Font(bold=True, size=11, color=DK_GREEN)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_cf[f'{col}13'] = f'={col}12+{col}10'
    ws_cf[f'{col}13'].number_format = cur_fmt

ws_cf.cell(row=14, column=1, value="")
ws_cf.cell(row=15, column=1, value="Monthly Burn Rate")
for i in range(5):
    col = get_column_letter(2 + i)
    ws_cf[f'{col}15'] = f"=('P&L'!{col}18+ABS({col}4))/12"
    ws_cf[f'{col}15'].number_format = cur_fmt

ws_cf.cell(row=16, column=1, value="Runway (Months)")
ws_cf['A16'].font = Font(bold=True)
for i in range(5):
    col = get_column_letter(2 + i)
    ws_cf[f'{col}16'] = f'=IF({col}15=0,0,{col}13/{col}15)'
    ws_cf[f'{col}16'].number_format = '0.0'

freeze(ws_cf)

# ── Sheet 8: Cap Table ────────────────────────────────────────────
ws_ct = wb.create_sheet("Cap Table")
ws_ct.sheet_properties.tabColor = GREY_TAB
set_col_widths(ws_ct, a_width=30, other_width=20, max_col=4)

ws_ct.cell(row=1, column=1, value="Cap Table").font = Font(bold=True, size=14, color=DK_GREEN)
style_header_row(ws_ct, 2, 4)
ws_ct.cell(row=2, column=1, value="")
ws_ct.cell(row=2, column=2, value="Ownership %")
ws_ct.cell(row=2, column=3, value="Amount ($)")
ws_ct.cell(row=2, column=4, value="Valuation ($)")

# Pre-seed
ws_ct.cell(row=3, column=1, value="Founders (Pre-Seed)").font = Font(bold=True)
ws_ct['B3'] = 1.0
ws_ct['B3'].number_format = pct_fmt

ws_ct.cell(row=4, column=1, value="Option Pool (carved out)")
ws_ct['B4'] = 0.10
ws_ct['B4'].number_format = pct_fmt

ws_ct.cell(row=5, column=1, value="Founders (Post-Pool)")
ws_ct['B5'] = '=B3-B4'
ws_ct['B5'].number_format = pct_fmt

# Seed
ws_ct.cell(row=6, column=1, value="")
ws_ct.cell(row=7, column=1, value="SEED ROUND").font = Font(bold=True, size=11, color=DK_GREEN)
style_header_row(ws_ct, 7, 4)
ws_ct.cell(row=7, column=2, value="Ownership %")
ws_ct.cell(row=7, column=3, value="Amount ($)")
ws_ct.cell(row=7, column=4, value="Valuation ($)")

ws_ct.cell(row=8, column=1, value="Amount Raised")
ws_ct['C8'] = '=Assumptions!B51'
ws_ct['C8'].number_format = cur_fmt

ws_ct.cell(row=9, column=1, value="Pre-money Valuation")
ws_ct['D9'] = '=Assumptions!C51'
ws_ct['D9'].number_format = cur_fmt

ws_ct.cell(row=10, column=1, value="Post-money Valuation")
ws_ct['D10'] = '=C8+D9'
ws_ct['D10'].number_format = cur_fmt

ws_ct.cell(row=11, column=1, value="Seed Investors %")
ws_ct['B11'] = '=C8/D10'
ws_ct['B11'].number_format = pct_fmt

ws_ct.cell(row=12, column=1, value="Founders post-Seed")
ws_ct['B12'] = '=(1-B4-B11)'
ws_ct['B12'].number_format = pct_fmt

ws_ct.cell(row=13, column=1, value="Option Pool post-Seed")
ws_ct['B13'] = '=B4'
ws_ct['B13'].number_format = pct_fmt

# Series A
ws_ct.cell(row=14, column=1, value="")
ws_ct.cell(row=15, column=1, value="SERIES A").font = Font(bold=True, size=11, color=DK_GREEN)
style_header_row(ws_ct, 15, 4)
ws_ct.cell(row=15, column=2, value="Ownership %")
ws_ct.cell(row=15, column=3, value="Amount ($)")
ws_ct.cell(row=15, column=4, value="Valuation ($)")

ws_ct.cell(row=16, column=1, value="Amount Raised")
ws_ct['C16'] = '=Assumptions!B52'
ws_ct['C16'].number_format = cur_fmt

ws_ct.cell(row=17, column=1, value="Pre-money Valuation")
ws_ct['D17'] = '=Assumptions!C52'
ws_ct['D17'].number_format = cur_fmt

ws_ct.cell(row=18, column=1, value="Post-money Valuation")
ws_ct['D18'] = '=C16+D17'
ws_ct['D18'].number_format = cur_fmt

ws_ct.cell(row=19, column=1, value="Series A Investors %")
ws_ct['B19'] = '=C16/D18'
ws_ct['B19'].number_format = pct_fmt

ws_ct.cell(row=20, column=1, value="Founders post-A (diluted)")
ws_ct['B20'] = '=B12*(1-B19)'
ws_ct['B20'].number_format = pct_fmt

ws_ct.cell(row=21, column=1, value="Seed Investors post-A (diluted)")
ws_ct['B21'] = '=B11*(1-B19)'
ws_ct['B21'].number_format = pct_fmt

ws_ct.cell(row=22, column=1, value="Option Pool post-A (diluted)")
ws_ct['B22'] = '=B13*(1-B19)'
ws_ct['B22'].number_format = pct_fmt

# Series B
ws_ct.cell(row=23, column=1, value="")
ws_ct.cell(row=24, column=1, value="SERIES B").font = Font(bold=True, size=11, color=DK_GREEN)
style_header_row(ws_ct, 24, 4)
ws_ct.cell(row=24, column=2, value="Ownership %")
ws_ct.cell(row=24, column=3, value="Amount ($)")
ws_ct.cell(row=24, column=4, value="Valuation ($)")

ws_ct.cell(row=25, column=1, value="Amount Raised")
ws_ct['C25'] = '=Assumptions!B53'
ws_ct['C25'].number_format = cur_fmt

ws_ct.cell(row=26, column=1, value="Pre-money Valuation")
ws_ct['D26'] = '=Assumptions!C53'
ws_ct['D26'].number_format = cur_fmt

ws_ct.cell(row=27, column=1, value="Post-money Valuation")
ws_ct['D27'] = '=C25+D26'
ws_ct['D27'].number_format = cur_fmt

ws_ct.cell(row=28, column=1, value="Series B Investors %")
ws_ct['B28'] = '=C25/D27'
ws_ct['B28'].number_format = pct_fmt

ws_ct.cell(row=29, column=1, value="Founders post-B (diluted)")
ws_ct['B29'] = '=B20*(1-B28)'
ws_ct['B29'].number_format = pct_fmt

ws_ct.cell(row=30, column=1, value="Seed Investors post-B (diluted)")
ws_ct['B30'] = '=B21*(1-B28)'
ws_ct['B30'].number_format = pct_fmt

ws_ct.cell(row=31, column=1, value="Series A Investors post-B (diluted)")
ws_ct['B31'] = '=B19*(1-B28)'
ws_ct['B31'].number_format = pct_fmt

ws_ct.cell(row=32, column=1, value="Option Pool post-B (diluted)")
ws_ct['B32'] = '=B22*(1-B28)'
ws_ct['B32'].number_format = pct_fmt

ws_ct.cell(row=33, column=1, value="Series B Investors")
ws_ct['B33'] = '=B28'
ws_ct['B33'].number_format = pct_fmt

ws_ct.cell(row=34, column=1, value="")
ws_ct.cell(row=35, column=1, value="TOTAL (verify = 100%)").font = Font(bold=True)
ws_ct['B35'] = '=B29+B30+B31+B32+B33'
ws_ct['B35'].number_format = pct_fmt

freeze(ws_ct, 'A3')

# ── Sheet 9: Charts ──────────────────────────────────────────────
ws_ch = wb.create_sheet("Charts")
ws_ch.sheet_properties.tabColor = GREEN_TAB

# -- Data tables for charts (hidden at the bottom) --
# Revenue by stream data
ch_data_start = 2
ws_ch.cell(row=ch_data_start, column=1, value="Year")
ws_ch.cell(row=ch_data_start, column=2, value="Hardware+Sub")
ws_ch.cell(row=ch_data_start, column=3, value="Data Licensing")
ws_ch.cell(row=ch_data_start, column=4, value="Weather")
ws_ch.cell(row=ch_data_start, column=5, value="Platform IP")
for i in range(5):
    r = ch_data_start + 1 + i
    col_src = get_column_letter(2 + i)
    ws_ch.cell(row=r, column=1, value=f"Y{i+1}")
    ws_ch[f'B{r}'] = f"='Revenue Buildup'!{col_src}6"
    ws_ch[f'B{r}'].number_format = cur_fmt
    ws_ch[f'C{r}'] = f"='Revenue Buildup'!{col_src}14"
    ws_ch[f'C{r}'].number_format = cur_fmt
    ws_ch[f'D{r}'] = f"='Revenue Buildup'!{col_src}20"
    ws_ch[f'D{r}'].number_format = cur_fmt
    ws_ch[f'E{r}'] = f"='Revenue Buildup'!{col_src}24"
    ws_ch[f'E{r}'].number_format = cur_fmt

# Chart 1: Revenue by Stream (Stacked Bar)
chart1 = BarChart()
chart1.type = "col"
chart1.grouping = "stacked"
chart1.title = "Revenue by Stream (Y1-Y5)"
chart1.y_axis.title = "Revenue ($)"
chart1.x_axis.title = "Year"
chart1.style = 10
chart1.width = 20
chart1.height = 14

cats = Reference(ws_ch, min_col=1, min_row=ch_data_start+1, max_row=ch_data_start+5)
for ci in range(2, 6):
    vals = Reference(ws_ch, min_col=ci, min_row=ch_data_start, max_row=ch_data_start+5)
    chart1.add_data(vals, titles_from_data=True)
chart1.set_categories(cats)
ws_ch.add_chart(chart1, "A9")

# -- Scenario comparison data --
sc_data_start = 10
ws_ch.cell(row=sc_data_start, column=7, value="Year")
ws_ch.cell(row=sc_data_start, column=8, value="Conservative")
ws_ch.cell(row=sc_data_start, column=9, value="Base")
ws_ch.cell(row=sc_data_start, column=10, value="Optimistic")
for i in range(5):
    r = sc_data_start + 1 + i
    ws_ch.cell(row=r, column=7, value=f"Y{i+1}")
    cons_col = get_column_letter(2 + i)  # Scenarios sheet Conservative
    base_col = get_column_letter(8 + i)
    opt_col = get_column_letter(14 + i)
    ws_ch[f'H{r}'] = f"=Scenarios!{cons_col}5"
    ws_ch[f'H{r}'].number_format = cur_fmt
    ws_ch[f'I{r}'] = f"=Scenarios!{base_col}5"
    ws_ch[f'I{r}'].number_format = cur_fmt
    ws_ch[f'J{r}'] = f"=Scenarios!{opt_col}5"
    ws_ch[f'J{r}'].number_format = cur_fmt

# Chart 2: Scenario Comparison (Line)
chart2 = LineChart()
chart2.title = "Scenario Comparison \u2014 Total Revenue"
chart2.y_axis.title = "Revenue ($)"
chart2.x_axis.title = "Year"
chart2.style = 10
chart2.width = 20
chart2.height = 14

cats2 = Reference(ws_ch, min_col=7, min_row=sc_data_start+1, max_row=sc_data_start+5)
for ci in range(8, 11):
    vals = Reference(ws_ch, min_col=ci, min_row=sc_data_start, max_row=sc_data_start+5)
    chart2.add_data(vals, titles_from_data=True)
chart2.set_categories(cats2)
ws_ch.add_chart(chart2, "L9")

# -- Gross Margin data --
gm_start = 18
ws_ch.cell(row=gm_start, column=1, value="Year")
ws_ch.cell(row=gm_start, column=2, value="Gross Margin %")
for i in range(5):
    r = gm_start + 1 + i
    col_src = get_column_letter(2 + i)
    ws_ch.cell(row=r, column=1, value=f"Y{i+1}")
    ws_ch[f'B{r}'] = f"='P&L'!{col_src}13"
    ws_ch[f'B{r}'].number_format = pct_fmt

# Chart 3: Gross Margin Evolution (Line)
chart3 = LineChart()
chart3.title = "Gross Margin Evolution"
chart3.y_axis.title = "Gross Margin %"
chart3.y_axis.numFmt = '0%'
chart3.style = 10
chart3.width = 20
chart3.height = 14

cats3 = Reference(ws_ch, min_col=1, min_row=gm_start+1, max_row=gm_start+5)
vals3 = Reference(ws_ch, min_col=2, min_row=gm_start, max_row=gm_start+5)
chart3.add_data(vals3, titles_from_data=True)
chart3.set_categories(cats3)
ws_ch.add_chart(chart3, "A27")

# -- Cash Position data --
cash_start = 25
ws_ch.cell(row=cash_start, column=7, value="Year")
ws_ch.cell(row=cash_start, column=8, value="Closing Cash")
for i in range(5):
    r = cash_start + 1 + i
    col_src = get_column_letter(2 + i)
    ws_ch.cell(row=r, column=7, value=f"Y{i+1}")
    ws_ch[f'H{r}'] = f"='Cash Flow'!{col_src}13"
    ws_ch[f'H{r}'].number_format = cur_fmt

# Chart 4: Cash Position (Bar)
chart4 = BarChart()
chart4.type = "col"
chart4.title = "Cash Position by Year"
chart4.y_axis.title = "Cash ($)"
chart4.style = 10
chart4.width = 20
chart4.height = 14

cats4 = Reference(ws_ch, min_col=7, min_row=cash_start+1, max_row=cash_start+5)
vals4 = Reference(ws_ch, min_col=8, min_row=cash_start, max_row=cash_start+5)
chart4.add_data(vals4, titles_from_data=True)
chart4.set_categories(cats4)
ws_ch.add_chart(chart4, "L27")

# -- Device Deployment data --
dev_start = 33
ws_ch.cell(row=dev_start, column=1, value="Year")
ws_ch.cell(row=dev_start, column=2, value="Cumulative Devices")
for i in range(5):
    r = dev_start + 1 + i
    col_src = get_column_letter(2 + i)
    ws_ch.cell(row=r, column=1, value=f"Y{i+1}")
    ws_ch[f'B{r}'] = f"=Assumptions!{col_src}12"
    ws_ch[f'B{r}'].number_format = num_fmt

# Chart 5: Device Deployment (Area)
chart5 = AreaChart()
chart5.title = "Device Deployment (Cumulative)"
chart5.y_axis.title = "Devices"
chart5.style = 10
chart5.width = 20
chart5.height = 14

cats5 = Reference(ws_ch, min_col=1, min_row=dev_start+1, max_row=dev_start+5)
vals5 = Reference(ws_ch, min_col=2, min_row=dev_start, max_row=dev_start+5)
chart5.add_data(vals5, titles_from_data=True)
chart5.set_categories(cats5)
ws_ch.add_chart(chart5, "A45")

# ── Save ──────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Financial model saved to: {OUT}")
